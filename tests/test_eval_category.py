"""
รัน Evaluation ทีละหมวด (เพื่อกัน rate limit)

ใช้:
    python -m tests.test_eval_category curriculum      ← รันหมวดเดียว
    python -m tests.test_eval_category staff
    python -m tests.test_eval_category --all           ← รันทุกหมวด (ใช้เวลานาน!)

Output ต่อหมวด:
    evaluation_results/{category}_{timestamp}.csv
    evaluation_results/{category}_{timestamp}.xlsx
"""

import sys
import time
import csv
import datetime
import re
from pathlib import Path

from src.config import PROJECT_ROOT
from src.chat import answer
from tests.eval_data import ALL_CATEGORIES, get_questions, get_category_th


# ─── Settings ─────────────────────────────────────
DELAY_BETWEEN_TESTS = 5  # วินาที
OUTPUT_DIR = PROJECT_ROOT / "evaluation_results"
OUTPUT_DIR.mkdir(exist_ok=True)


# ═══════════════════════════════════════════════════════
# 4 Metrics
# ═══════════════════════════════════════════════════════

def measure_answer_relevance(response, expected_keywords, alt_keywords):
    """Metric 1: Answer Relevance"""
    response_lower = response.lower()
    
    missing_expected = [kw for kw in expected_keywords if kw.lower() not in response_lower]
    has_all_expected = len(missing_expected) == 0
    
    if alt_keywords:
        matching_alt = [kw for kw in alt_keywords if kw.lower() in response_lower]
        has_any_alt = len(matching_alt) > 0
    else:
        matching_alt = []
        has_any_alt = True
    
    if has_all_expected and has_any_alt:
        score = 1.0
    elif has_all_expected or (has_any_alt and not expected_keywords):
        score = 1.0
    elif has_all_expected:
        score = 0.7
    elif has_any_alt:
        score = 0.5
    else:
        score = 0.0
    
    return score, missing_expected


def measure_faithfulness(response, ground_truth):
    """Metric 2: Faithfulness — word overlap"""
    if not response or not ground_truth:
        return 0.0
    
    words_response = set(re.findall(r'\b\w{2,}\b', response.lower()))
    words_truth = set(re.findall(r'\b\w{2,}\b', ground_truth.lower()))
    
    if not words_truth:
        return 0.0
    
    intersection = words_response & words_truth
    union = words_response | words_truth
    return len(intersection) / len(union) if union else 0.0


def measure_oos_rejection(response, is_oos):
    """Metric 3: OOS Rejection"""
    rejection_signals = [
        "ไม่มีข้อมูล", "ไม่พบ", "ไม่ทราบ",
        "ขออภัย", "ติดต่อภาควิชา",
        "ไม่สามารถตอบ", "นอกเหนือ",
    ]
    rejected = any(sig in response for sig in rejection_signals)
    
    if is_oos:
        return (1.0, "✅ Correctly rejected") if rejected else (0.0, "❌ Failed to reject OOS")
    else:
        return (0.0, "❌ Incorrectly rejected") if rejected else (1.0, "✅ Correctly answered")


def measure_latency(start, end):
    """Metric 4: Latency"""
    return round(end - start, 3)


# ═══════════════════════════════════════════════════════
# Test Runner
# ═══════════════════════════════════════════════════════

def evaluate_single(case, total_count, current_idx):
    """ทดสอบ 1 คำถาม"""
    question = case["question"]
    print(f"\n[{current_idx:02d}/{total_count}] {question[:70]}")
    
    start = time.time()
    try:
        response = answer(question, verbose=False)
        end = time.time()
        error = None
    except Exception as e:
        end = time.time()
        response = ""
        error = str(e)[:200]
    
    latency = measure_latency(start, end)
    
    if error:
        print(f"   ❌ ERROR: {error[:100]}")
        return {
            **case, "response": "", "error": error, "latency": latency,
            "answer_relevance": 0.0, "faithfulness": 0.0,
            "oos_score": 0.0, "oos_verdict": "ERROR", "passed": False,
            "missing_expected": [],
        }
    
    relevance, missing = measure_answer_relevance(
        response, case["expected_keywords"], case["alt_keywords"]
    )
    faithfulness = measure_faithfulness(response, case["ground_truth"])
    oos_score, oos_verdict = measure_oos_rejection(response, case["is_oos"])
    
    passed = (relevance >= 0.7) and (oos_score == 1.0)
    
    icon = "✅" if passed else "❌"
    print(f"   {icon} Rel:{relevance:.2f} | Faith:{faithfulness:.2f} | Latency:{latency:.1f}s")
    print(f"   💬 {response[:120]}")
    
    return {
        **case, "response": response, "error": None, "latency": latency,
        "answer_relevance": round(relevance, 3),
        "faithfulness": round(faithfulness, 3),
        "oos_score": oos_score, "oos_verdict": oos_verdict,
        "passed": passed, "missing_expected": missing,
    }


def run_category(category):
    """รัน evaluation 1 หมวด"""
    if category not in ALL_CATEGORIES:
        print(f"❌ Unknown category: {category}")
        print(f"   Available: {list(ALL_CATEGORIES.keys())}")
        return None
    
    questions = get_questions(category)
    category_th = get_category_th(category)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    print("=" * 70)
    print(f"🧪 Evaluating: {category_th} ({category})")
    print("=" * 70)
    print(f"📊 Questions: {len(questions)}")
    print(f"⏱️  Delay: {DELAY_BETWEEN_TESTS}s")
    print(f"⏳ Estimated: ~{len(questions) * (DELAY_BETWEEN_TESTS + 3) / 60:.1f} minutes")
    print("=" * 70)
    
    results = []
    for i, case in enumerate(questions, 1):
        result = evaluate_single(case, len(questions), i)
        results.append(result)
        if i < len(questions):
            time.sleep(DELAY_BETWEEN_TESTS)
    
    print_summary(results, category_th)
    save_csv(results, category, timestamp)
    save_excel(results, category, category_th, timestamp)
    
    print(f"\n📁 Saved to: {OUTPUT_DIR}/")
    print(f"   - {category}_{timestamp}.csv")
    print(f"   - {category}_{timestamp}.xlsx")
    
    return results


def print_summary(results, category_th):
    """พิมพ์สรุป"""
    total = len(results)
    if total == 0:
        return
    
    passed = sum(1 for r in results if r["passed"])
    errors = sum(1 for r in results if r["error"])
    avg_rel = sum(r["answer_relevance"] for r in results) / total
    avg_faith = sum(r["faithfulness"] for r in results) / total
    avg_oos = sum(r["oos_score"] for r in results) / total
    avg_lat = sum(r["latency"] for r in results) / total
    
    print()
    print("=" * 70)
    print(f"📊 Summary: {category_th}")
    print("=" * 70)
    print(f"   Total:    {total}")
    print(f"   Passed:   {passed} ({passed/total*100:.1f}%)")
    print(f"   Failed:   {total - passed - errors}")
    print(f"   Errors:   {errors}")
    print(f"\n   Avg Answer Relevance: {avg_rel:.3f}")
    print(f"   Avg Faithfulness:     {avg_faith:.3f}")
    print(f"   Avg OOS Score:        {avg_oos:.3f}")
    print(f"   Avg Latency:          {avg_lat:.2f} sec")
    print("=" * 70)


def save_csv(results, category, timestamp):
    """บันทึก CSV"""
    if not results:
        return
    filepath = OUTPUT_DIR / f"{category}_{timestamp}.csv"
    
    fieldnames = [
        "id", "category", "question", "ground_truth", "response",
        "expected_keywords", "alt_keywords", "missing_expected",
        "answer_relevance", "faithfulness", "oos_score", "oos_verdict",
        "latency", "passed", "is_oos", "error",
    ]
    
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for r in results:
            row = {**r}
            row["expected_keywords"] = ", ".join(r["expected_keywords"])
            row["alt_keywords"] = ", ".join(r["alt_keywords"])
            row["missing_expected"] = ", ".join(r.get("missing_expected", []))
            writer.writerow(row)


def save_excel(results, category, category_th, timestamp):
    """บันทึก Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("⚠️  ติดตั้ง openpyxl เพื่อสร้าง Excel: pip install openpyxl")
        return
    
    if not results:
        return
    
    filepath = OUTPUT_DIR / f"{category}_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    
    # ─── Summary sheet ───
    ws = wb.active
    ws.title = "Summary"
    
    total = len(results)
    if total == 0:
        wb.save(filepath)
        return
    
    passed = sum(1 for r in results if r["passed"])
    
    ws["A1"] = f"Evaluation: {category_th}"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    overall = [
        ["Metric", "Value"],
        ["Total Questions", total],
        ["Passed", passed],
        ["Pass Rate", f"{passed/total*100:.1f}%"],
        ["Avg Answer Relevance", round(sum(r["answer_relevance"] for r in results) / total, 3)],
        ["Avg Faithfulness", round(sum(r["faithfulness"] for r in results) / total, 3)],
        ["Avg OOS Score", round(sum(r["oos_score"] for r in results) / total, 3)],
        ["Avg Latency (sec)", round(sum(r["latency"] for r in results) / total, 2)],
    ]
    
    for i, row in enumerate(overall, start=4):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
        if i == 4:
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.cell(row=i, column=2).font = Font(bold=True)
    
    # ─── Details sheet ───
    ws2 = wb.create_sheet("Details")
    headers = ["ID", "Question", "Response", "Relevance", "Faithfulness", "OOS", "Latency", "Passed"]
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    
    for i, r in enumerate(results, start=2):
        ws2.cell(row=i, column=1, value=r["id"])
        ws2.cell(row=i, column=2, value=r["question"])
        ws2.cell(row=i, column=3, value=r["response"][:200])
        ws2.cell(row=i, column=4, value=r["answer_relevance"])
        ws2.cell(row=i, column=5, value=r["faithfulness"])
        ws2.cell(row=i, column=6, value=r["oos_score"])
        ws2.cell(row=i, column=7, value=r["latency"])
        cell = ws2.cell(row=i, column=8, value="PASS" if r["passed"] else "FAIL")
        cell.fill = PatternFill("solid", fgColor="C6EFCE" if r["passed"] else "FFC7CE")
    
    # Column widths
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 40
    ws2.column_dimensions["C"].width = 60
    
    wb.save(filepath)


def main():
    if len(sys.argv) < 2:
        print("Usage: python -m tests.test_eval_category <category>")
        print(f"\nAvailable categories: {list(ALL_CATEGORIES.keys())}")
        print("   หรือ --all เพื่อรันทุกหมวด")
        return
    
    arg = sys.argv[1]
    
    if arg == "--all":
        print("⚠️  รันทุกหมวด 7 หมวด × 20 คำถาม = 140 คำถาม")
        print(f"⏳ ใช้เวลา ~{140 * (DELAY_BETWEEN_TESTS + 3) / 60:.0f} นาที")
        confirm = input("ยืนยัน? (y/n): ")
        if confirm.lower() != "y":
            return
        
        for cat in ALL_CATEGORIES.keys():
            print(f"\n\n{'#'*70}")
            print(f"# Starting: {cat}")
            print(f"{'#'*70}")
            run_category(cat)
            print(f"\n⏸️  รอ 30 วินาทีก่อนหมวดถัดไป...")
            time.sleep(30)
    else:
        run_category(arg)


if __name__ == "__main__":
    main()