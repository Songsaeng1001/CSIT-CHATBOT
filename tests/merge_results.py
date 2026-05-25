"""
รวมผลทุกหมวดเป็น Excel + Report เดียว

ใช้: python -m tests.merge_results

จะอ่าน CSV ล่าสุดของแต่ละหมวดจาก evaluation_results/
แล้วรวมเป็น combined Excel
"""

import csv
import datetime
from pathlib import Path
from collections import defaultdict

from src.config import PROJECT_ROOT
from tests.eval_data import ALL_CATEGORIES, get_category_th


OUTPUT_DIR = PROJECT_ROOT / "evaluation_results"

def safe_sheet_title(title, max_len=31):
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for ch in invalid_chars:
        title = title.replace(ch, '-')
    return title[:max_len]


def find_latest_csv(category):
    """หา CSV ล่าสุดของหมวด"""
    files = sorted(OUTPUT_DIR.glob(f"{category}_*.csv"))
    return files[-1] if files else None


def load_csv(path):
    """อ่าน CSV"""
    results = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row["id"] = int(row["id"])
            row["answer_relevance"] = float(row["answer_relevance"])
            row["faithfulness"] = float(row["faithfulness"])
            row["oos_score"] = float(row["oos_score"])
            row["latency"] = float(row["latency"])
            row["passed"] = row["passed"].lower() == "true"
            results.append(row)
    return results


def merge_all():
    """รวมผลทุกหมวด"""
    all_results = {}
    
    for category in ALL_CATEGORIES.keys():
        csv_path = find_latest_csv(category)
        if csv_path:
            all_results[category] = load_csv(csv_path)
            print(f"✅ {category:15s}: {len(all_results[category])} results from {csv_path.name}")
        else:
            print(f"⚠️  {category:15s}: ไม่พบ CSV")
    
    if not all_results:
        print("❌ ไม่พบผลทดสอบ — รัน python -m tests.test_eval_category <category> ก่อน")
        return
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    save_combined_excel(all_results, timestamp)
    save_combined_report(all_results, timestamp)


def save_combined_excel(all_results, timestamp):
    """สร้าง Excel รวม"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.chart import BarChart, Reference
    except ImportError:
        print("⚠️  ติดตั้ง openpyxl: pip install openpyxl")
        return
    
    filepath = OUTPUT_DIR / f"combined_{timestamp}.xlsx"
    wb = openpyxl.Workbook()
    
    # ─── Summary sheet ───
    ws = wb.active
    ws.title = "Overall Summary"
    
    ws["A1"] = "น้องซีที — Combined Evaluation Summary"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Stats overall
    total_questions = sum(len(r) for r in all_results.values())
    total_passed = sum(sum(1 for q in r if q["passed"]) for r in all_results.values())
    
    ws["A4"] = "Overall Metrics"
    ws["A4"].font = Font(bold=True, size=12)
    
    overall_avg_rel = sum(q["answer_relevance"] for r in all_results.values() for q in r) / total_questions if total_questions else 0
    overall_avg_faith = sum(q["faithfulness"] for r in all_results.values() for q in r) / total_questions if total_questions else 0
    overall_avg_oos = sum(q["oos_score"] for r in all_results.values() for q in r) / total_questions if total_questions else 0
    overall_avg_lat = sum(q["latency"] for r in all_results.values() for q in r) / total_questions if total_questions else 0
    
    overall = [
        ["Total Questions", total_questions],
        ["Total Passed", total_passed],
        ["Overall Pass Rate", f"{total_passed/total_questions*100:.1f}%" if total_questions else "N/A"],
        ["Avg Answer Relevance", round(overall_avg_rel, 3)],
        ["Avg Faithfulness", round(overall_avg_faith, 3)],
        ["Avg OOS Score", round(overall_avg_oos, 3)],
        ["Avg Latency (sec)", round(overall_avg_lat, 2)],
    ]
    
    for i, (key, val) in enumerate(overall, start=5):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=val)
    
    # ─── By category table ───
    ws["A14"] = "By Category"
    ws["A14"].font = Font(bold=True, size=12)
    
    headers = ["Category", "Total", "Passed", "Pass %", "Avg Rel", "Avg Faith", "Avg OOS", "Avg Latency"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=15, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    
    row_idx = 16
    for category, results in all_results.items():
        if not results:
            continue
        n = len(results)
        p = sum(1 for r in results if r["passed"])
        
        ws.cell(row=row_idx, column=1, value=get_category_th(category))
        ws.cell(row=row_idx, column=2, value=n)
        ws.cell(row=row_idx, column=3, value=p)
        ws.cell(row=row_idx, column=4, value=round(p/n*100, 1))
        ws.cell(row=row_idx, column=5, value=round(sum(r["answer_relevance"] for r in results) / n, 3))
        ws.cell(row=row_idx, column=6, value=round(sum(r["faithfulness"] for r in results) / n, 3))
        ws.cell(row=row_idx, column=7, value=round(sum(r["oos_score"] for r in results) / n, 3))
        ws.cell(row=row_idx, column=8, value=round(sum(r["latency"] for r in results) / n, 2))
        row_idx += 1
    
    # ─── Add chart ───
    chart = BarChart()
    chart.title = "Pass Rate by Category"
    chart.y_axis.title = "Pass Rate (%)"
    
    data = Reference(ws, min_col=4, min_row=15, max_row=row_idx-1, max_col=4)
    categories = Reference(ws, min_col=1, min_row=16, max_row=row_idx-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    
    ws.add_chart(chart, "J5")
    
    # ─── Individual category sheets ───
    for category, results in all_results.items():
        ws_cat = wb.create_sheet(safe_sheet_title(get_category_th(category), 15)) # max 15 chars
        
        headers2 = ["ID", "Question", "Response", "Rel", "Faith", "OOS", "Latency", "Pass"]
        for col, h in enumerate(headers2, start=1):
            cell = ws_cat.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
        
        for i, r in enumerate(results, start=2):
            ws_cat.cell(row=i, column=1, value=r["id"])
            ws_cat.cell(row=i, column=2, value=r["question"])
            ws_cat.cell(row=i, column=3, value=r["response"][:200])
            ws_cat.cell(row=i, column=4, value=r["answer_relevance"])
            ws_cat.cell(row=i, column=5, value=r["faithfulness"])
            ws_cat.cell(row=i, column=6, value=r["oos_score"])
            ws_cat.cell(row=i, column=7, value=r["latency"])
            cell = ws_cat.cell(row=i, column=8, value="PASS" if r["passed"] else "FAIL")
            cell.fill = PatternFill("solid", fgColor="C6EFCE" if r["passed"] else "FFC7CE")
    
    wb.save(filepath)
    print(f"\n✅ Combined Excel: {filepath.name}")


def save_combined_report(all_results, timestamp):
    """สร้าง Markdown report รวม"""
    filepath = OUTPUT_DIR / f"combined_report_{timestamp}.md"
    
    total = sum(len(r) for r in all_results.values())
    passed = sum(sum(1 for q in r if q["passed"]) for r in all_results.values())
    
    if total == 0:
        return
    
    content = f"""# 📊 น้องซีที — Combined Evaluation Report

**Date:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Total Questions:** {total}  
**Categories:** {len(all_results)}

---

## 🎯 Overall Metrics

| Metric | Value |
|---|---|
| Total Questions | {total} |
| Passed | {passed} ({passed/total*100:.1f}%) |
| Failed | {total - passed} |
| Avg Answer Relevance | {sum(q['answer_relevance'] for r in all_results.values() for q in r) / total:.3f} |
| Avg Faithfulness | {sum(q['faithfulness'] for r in all_results.values() for q in r) / total:.3f} |
| Avg OOS Score | {sum(q['oos_score'] for r in all_results.values() for q in r) / total:.3f} |
| Avg Latency (sec) | {sum(q['latency'] for r in all_results.values() for q in r) / total:.2f} |

---

## 📂 By Category

| Category | Total | Passed | Pass Rate | Avg Rel | Avg Latency |
|---|---|---|---|---|---|
"""
    
    for category, results in all_results.items():
        n = len(results)
        if n == 0:
            continue
        p = sum(1 for r in results if r["passed"])
        avg_r = sum(r["answer_relevance"] for r in results) / n
        avg_l = sum(r["latency"] for r in results) / n
        content += f"| {get_category_th(category)} | {n} | {p} | {p/n*100:.1f}% | {avg_r:.3f} | {avg_l:.2f}s |\n"
    
    content += "\n---\n\n## ❌ Failed Cases\n"
    
    for category, results in all_results.items():
        failed = [r for r in results if not r["passed"]]
        if not failed:
            continue
        content += f"\n### {get_category_th(category)} ({len(failed)} failed)\n\n"
        for r in failed:
            content += f"- **Q{r['id']}**: {r['question']}\n"
            content += f"  - Response: {r['response'][:150]}\n"
            content += f"  - Relevance: {r['answer_relevance']:.2f}\n\n"
    
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)
    
    print(f"✅ Combined Report: {filepath.name}")


if __name__ == "__main__":
    merge_all()